import importlib.util
import shutil
import sys
import unittest
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_fixtures import (  # noqa: E402
    PRODUCTION_ROOT,
    TEST_DATE,
    ProductionDataGuardMixin,
    make_sandbox,
    write_canonical_states,
)

ROOT = PRODUCTION_ROOT
WORKFLOW_PATH = ROOT / "tools" / "workflow.py"
SPEC = importlib.util.spec_from_file_location("personal_workflow", WORKFLOW_PATH)
workflow = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = workflow
SPEC.loader.exec_module(workflow)

TICKETS_PATH = ROOT / "tools" / "common" / "tickets.py"
TSPEC = importlib.util.spec_from_file_location("tickets", TICKETS_PATH)
tickets = importlib.util.module_from_spec(TSPEC)
assert TSPEC.loader is not None
sys.modules[TSPEC.name] = tickets
TSPEC.loader.exec_module(tickets)

SPORTS = ("jra", "nar", "kyotei")
EXCEL_KEYS = (
    "jra_entry",
    "jra_summary",
    "nar_entry",
    "nar_summary",
    "kyotei_entry",
    "kyotei_summary",
)


class PersonalWorkflowTest(ProductionDataGuardMixin, unittest.TestCase):
    """通し確認は一時ディレクトリ＋ allow_sample=True（テストデータ使用）。"""

    def setUp(self):
        super().setUp()
        self.sandbox = make_sandbox(ROOT, copy_excel=True)
        self.addCleanup(shutil.rmtree, self.sandbox, True)
        write_canonical_states(self.sandbox, start_date=TEST_DATE)
        self._orig_root = workflow.ROOT
        workflow.ROOT = self.sandbox
        self.addCleanup(setattr, workflow, "ROOT", self._orig_root)

    def _predict(self, sport: str, **kwargs):
        kwargs.setdefault("force", True)
        kwargs.setdefault("sync_drive", False)
        kwargs.setdefault("allow_sample", True)
        kwargs.setdefault("try_auto", False)
        return workflow.run_predict(sport, TEST_DATE, **kwargs)

    def _pred_doc(self, sport: str):
        from common.daily_json import load_predictions_doc, prediction_reread_problems

        doc = load_predictions_doc(workflow.ROOT, sport, TEST_DATE)
        self.assertIsNotNone(doc)
        self.assertEqual(prediction_reread_problems(doc, doc), [])
        return doc

    def _pred_races(self, sport: str):
        from common.daily_json import records_from_predictions_doc

        return records_from_predictions_doc(self._pred_doc(sport))

    def _result_doc(self, sport: str):
        from common.daily_json import load_results_doc, results_reread_problems

        doc = load_results_doc(workflow.ROOT, sport, TEST_DATE)
        self.assertIsNotNone(doc)
        self.assertEqual(results_reread_problems(doc, doc), [])
        return doc

    def _assert_state_records_empty(self, sport: str):
        state = workflow.load_json(workflow.state_path(sport))
        self.assertEqual(state.get("records"), [])

    def test_expand_pick(self):
        self.assertEqual(
            tickets.expand_pick("4-2-135"),
            ("4-2-1", "4-2-3", "4-2-5"),
        )

    def test_init_excel_six_files_and_months(self):
        msg = workflow.init_excel_cmd()
        self.assertIn("中央競馬", msg)
        self.assertIn("地方競馬", msg)
        self.assertIn("競艇", msg)
        excel = workflow.ensure_workbooks(workflow.ROOT)
        for key in EXCEL_KEYS:
            self.assertIn(key, excel)
            self.assertTrue(excel[key].exists())
        self.assertNotIn("keiba_entry", excel)
        self.assertNotIn("keirin_entry", excel)
        from openpyxl import load_workbook
        from common.constants import MONTH_SHEETS

        for path in excel.values():
            wb = load_workbook(path, read_only=True)
            for month in MONTH_SHEETS:
                self.assertIn(month, wb.sheetnames, f"{path.name} に {month} がない")
            wb.close()

    def test_predict_jra_from_sample(self):
        """テストデータ使用: allow_sample=True で中央競馬の通しを確認する。"""
        report = self._predict("jra")
        self.assertIn("中央競馬", report)
        self.assertIn("中山", report)
        self.assertIn("学習JSON保存", report)
        doc = self._pred_doc("jra")
        self.assertEqual(doc["date"], TEST_DATE)
        self.assertEqual(doc["sport"], "jra")
        selected = self._pred_races("jra")
        self.assertGreaterEqual(len(selected), 1)
        self.assertLessEqual(len(selected), 5)
        self.assertTrue(all(r.get("sport") == "jra" for r in selected))
        for race in selected:
            self.assertTrue(race.get("race_id"))
            self.assertIn("prediction_score", race)
            self.assertIn("confidence", race)
            self.assertTrue(race.get("tickets"))
            self.assertGreaterEqual(race.get("ticket_count") or 0, 1)
            self.assertTrue(race.get("axis") or (race.get("tickets") or [{}])[0].get("pick"))
        self._assert_state_records_empty("jra")

    def test_predict_nar_max_five(self):
        """テストデータ使用: 地方競馬は最大5レース。"""
        report = self._predict("nar")
        self.assertIn("地方競馬", report)
        self.assertIn("大井", report)
        self.assertIn("学習JSON保存", report)
        selected = self._pred_races("nar")
        self.assertEqual(len(selected), 5)
        self.assertTrue(all(r.get("sport") == "nar" for r in selected))
        venues = {r["venue"] for r in selected}
        self.assertNotIn("中山", venues)
        self._assert_state_records_empty("nar")

    def test_predict_kyotei_from_sample(self):
        """テストデータ使用: 競艇の通しを確認する。"""
        report = self._predict("kyotei")
        self.assertIn("競艇", report)
        self.assertIn("学習JSON保存", report)
        selected = self._pred_races("kyotei")
        self.assertGreaterEqual(len(selected), 1)
        self.assertLessEqual(len(selected), 5)
        self._assert_state_records_empty("kyotei")

    def test_predict_keiba_and_keirin_not_used(self):
        self.assertIn("未対応", workflow.run_predict("keiba", TEST_DATE, force=True, sync_drive=False))
        self.assertIn("未対応", workflow.run_predict("keirin", TEST_DATE, force=True, sync_drive=False))

    def test_learning_data_is_separated(self):
        self._predict("jra")
        self._predict("nar")
        workflow.apply_results_from_file(
            "jra", TEST_DATE, ROOT / "examples" / "jra_results.sample.json", sync_drive=False
        )
        workflow.apply_results_from_file(
            "nar", TEST_DATE, ROOT / "examples" / "nar_results.sample.json", sync_drive=False
        )
        jra_venues = {r.get("venue") for r in self._pred_races("jra")}
        nar_venues = {r.get("venue") for r in self._pred_races("nar")}
        self.assertTrue(jra_venues)
        self.assertTrue(nar_venues)
        self.assertFalse(jra_venues & nar_venues)
        self._assert_state_records_empty("jra")
        self._assert_state_records_empty("nar")
        workflow.ingest_inbox("jra", TEST_DATE)
        workflow.ingest_inbox("nar", TEST_DATE)
        jra_learn = workflow.run_learning_report("jra")
        nar_learn = workflow.run_learning_report("nar")
        self.assertIn("100レースまでの残り", jra_learn)
        self.assertIn("100レースまでの残り", nar_learn)
        self.assertTrue((workflow.ROOT / "data" / "jra" / "learning_report.json").exists())
        self.assertTrue((workflow.ROOT / "data" / "nar" / "learning_report.json").exists())
        self.assertNotEqual(
            workflow.state_path("jra"),
            workflow.state_path("nar"),
        )

    def test_idempotent_predict_without_force(self):
        self._predict("jra")
        second = self._predict("jra", force=False)
        self.assertIn("二重登録防止", second)

    def test_apply_results_and_review(self):
        self._predict("jra")
        report = workflow.apply_results_from_file(
            "jra", TEST_DATE, ROOT / "examples" / "jra_results.sample.json", sync_drive=False
        )
        self.assertIn("結果報告", report)
        self.assertIn("学習JSON保存", report)
        races = self._result_doc("jra")["races"]
        self.assertGreater(len(races), 0)
        for race in races:
            self.assertTrue(race.get("race_id"))
            self.assertIn(race.get("status"), {"的中", "ハズレ"})
            self.assertIn("stake", race)
            self.assertIn("payout", race)
            self.assertTrue(race.get("trifecta"))
            if race.get("status") == "ハズレ":
                self.assertTrue(race.get("primary_miss_reason"))
        self._assert_state_records_empty("jra")

    def test_partial_results_are_not_marked_processed(self):
        from common.daily_json import results_cover_predictions
        from fetch.race_builder import save_results_json

        self._predict("nar")
        selected = self._pred_races("nar")
        self.assertGreaterEqual(len(selected), 2)
        first, rest = selected[0], selected[1:]

        save_results_json(
            workflow.ROOT,
            "nar",
            TEST_DATE,
            [
                {
                    "venue": first["venue"],
                    "race": first["race"],
                    "trifecta": "1-2-3",
                    "payout": 1200,
                }
            ],
            source="auto_fetch",
        )
        with patch("orchestrator.fetch_keiba_results", return_value=[]):
            report = workflow.run_results("nar", TEST_DATE, force=True, sync_drive=False)
        self.assertIn("処理済みにはしません", report)
        pred_doc = self._pred_doc("nar")
        result_doc = self._result_doc("nar")
        self.assertFalse(results_cover_predictions(pred_doc, result_doc))
        with_result = result_doc["races"]
        self.assertEqual(len(with_result), 1)
        self.assertEqual(with_result[0]["venue"], first["venue"])
        self._assert_state_records_empty("nar")

        remaining_payload = [
            {
                "venue": r["venue"],
                "race": r["race"],
                "trifecta": "2-3-4",
                "payout": 800,
            }
            for r in rest
        ]
        with patch("orchestrator.fetch_keiba_results", return_value=remaining_payload):
            second = workflow.run_results("nar", TEST_DATE, force=True, sync_drive=False)
        self.assertNotIn("処理済みにはしません", second)
        result_doc = self._result_doc("nar")
        self.assertTrue(results_cover_predictions(self._pred_doc("nar"), result_doc))
        again = workflow.run_results("nar", TEST_DATE, force=False, sync_drive=False)
        self.assertIn("二重登録防止", again)
        first_after = next(
            r
            for r in result_doc["races"]
            if r.get("venue") == first["venue"] and r.get("race") == first["race"]
        )
        self.assertEqual(first_after["trifecta"], "1-2-3")

    def test_same_name_predictions_json_is_updated_not_duplicated(self):
        self._predict("jra")
        self._predict("jra", force=True)
        folder = workflow.ROOT / "data" / "inbox" / "jra"
        files = list(folder.glob("*.predictions.json"))
        self.assertEqual(len(files), 1)
        self.assertEqual(files[0].name, f"{TEST_DATE}.predictions.json")

    def test_learning_json_failure_does_not_undo_excel(self):
        from common.constants import EXCEL_FILENAMES
        from openpyxl import load_workbook

        excel_path = workflow.ROOT / "excel" / EXCEL_FILENAMES["jra_entry"]
        before = excel_path.read_bytes()
        with patch.object(
            workflow, "save_daily_json", side_effect=OSError("disk full")
        ):
            report = self._predict("jra")
        self.assertIn("学習JSON未保存", report)
        self.assertIn("中山", report)
        self.assertNotEqual(excel_path.read_bytes(), before)
        inbox = workflow.ROOT / "data" / "inbox" / "jra" / f"{TEST_DATE}.predictions.json"
        self.assertFalse(inbox.exists())
        self._assert_state_records_empty("jra")


if __name__ == "__main__":
    unittest.main()
