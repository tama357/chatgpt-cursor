"""第1段階（軸・予想しやすさ・外れ型の自動記録）に関する単体テスト。

既存A〜N列の処理・配点・閾値・買い目生成には触れない。新規追加した
excel/io.py・excel/mapping.py・common/constants.pyの追加分のみを対象にする。
"""

from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "tools"))

from common.constants import MISS_REASONS, MISS_TYPE_MAP, MISS_TYPE_NONE  # noqa: E402
from excel.io import _miss_type_label, write_predictions, write_results  # noqa: E402
from excel.mapping import inspect_entry_sheet, load_entry_mapping  # noqa: E402

ENTRY_HEADERS = [
    "日付",
    "予想番号",
    "狙い",
    "自信度",
    "競馬場",
    "R",
    "締切時刻",
    "本線",
    "抑え",
    "合計点数",
    "解説",
    "結果 3連単",
    "払戻金",
    "結果",
]
NEW_HEADERS = ["軸", "予想しやすさ", "外れ型"]


def _make_entry_workbook(path: Path, *, with_new_columns: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "202609"
    headers = list(ENTRY_HEADERS) + (NEW_HEADERS if with_new_columns else [])
    for col, title in enumerate(headers, start=1):
        ws.cell(2, col, title)
    for i in range(5):
        ws.cell(3 + i, 2, i + 1)  # 予想番号 1〜5
    wb.save(path)
    wb.close()


class MissTypeLabelTest(unittest.TestCase):
    """外れ型は14種類を1対1で日本語化する（4区分への圧縮はしない）。"""

    def test_all_14_reasons_map_1to1(self):
        self.assertEqual(len(MISS_TYPE_MAP), 14)
        self.assertEqual(set(MISS_TYPE_MAP.keys()), set(MISS_REASONS))
        self.assertEqual(len(set(MISS_TYPE_MAP.values())), 14)

    def test_hit_and_pending_are_dash(self):
        self.assertEqual(_miss_type_label({"status": "的中"}), MISS_TYPE_NONE)
        self.assertEqual(_miss_type_label({"status": "未実施"}), MISS_TYPE_NONE)

    def test_each_reason_label_matches_spec(self):
        expected = {
            "axis_miss": "軸外れ",
            "second_place_miss": "2着外れ",
            "third_place_miss": "3着外れ",
            "scenario_miss": "展開読み外れ",
            "line_collapse": "ライン崩壊",
            "unexpected_position": "想定外位置取り",
            "upset": "波乱",
            "condition_miss": "条件読み外れ",
            "accident": "事故",
            "data_shortage": "データ不足",
            "overconfidence": "過信",
            "too_many_combinations": "買い目過多",
            "too_few_combinations": "買い目不足",
            "other": "その他",
        }
        self.assertEqual(MISS_TYPE_MAP, expected)
        for reason, label in expected.items():
            result = {"status": "ハズレ", "primary_miss_reason": reason}
            self.assertEqual(_miss_type_label(result), label)

    def test_unknown_reason_falls_back_to_dash(self):
        result = {"status": "ハズレ", "primary_miss_reason": "not_a_real_reason"}
        self.assertEqual(_miss_type_label(result), MISS_TYPE_NONE)

    def test_secondary_miss_reasons_not_referenced(self):
        """secondary_miss_reasonsはExcelに出さない＝ここでは使わないことを確認。"""
        result = {
            "status": "ハズレ",
            "primary_miss_reason": "second_place_miss",
            "secondary_miss_reasons": ["scenario_miss", "too_few_combinations"],
        }
        self.assertEqual(_miss_type_label(result), "2着外れ")


class WritePredictionsNewColumnsTest(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="io-new-columns-"))
        self.addCleanup(shutil.rmtree, self.tmp, True)

    def _predictions(self):
        return [
            {
                "number": 1,
                "target": "鉄板",
                "confidence": "A",
                "venue": "中山",
                "race": 11,
                "close_time": "15:40",
                "tickets": [{"type": "本線", "pick": "1-2-3"}],
                "axis": "1",
                "prediction_score": 88,
            }
        ]

    def test_axis_and_prediction_score_written_when_columns_exist(self):
        path = self.tmp / "entry_with_columns.xlsx"
        _make_entry_workbook(path, with_new_columns=True)
        write_predictions(path, "2026-09-03", self._predictions())

        mapping = load_entry_mapping(path, "202609")
        row = mapping.row_for(3, 1)
        wb = load_workbook(path, data_only=True)
        ws = wb["202609"]
        self.assertEqual(ws.cell(row, mapping.columns["axis"]).value, "1")
        self.assertEqual(ws.cell(row, mapping.columns["prediction_score"]).value, 88)
        # 既存A〜N列相当（venue等）は今回の変更後も書かれていること
        self.assertEqual(ws.cell(row, mapping.columns["venue"]).value, "中山")
        self.assertEqual(ws.cell(row, mapping.columns["target"]).value, "鉄板")
        wb.close()

    def test_no_crash_and_existing_columns_unaffected_when_new_columns_absent(self):
        """列追加前の既存シートに対しても例外が出ず、既存列は変わらないこと。"""
        path = self.tmp / "entry_without_columns.xlsx"
        _make_entry_workbook(path, with_new_columns=False)
        write_predictions(path, "2026-09-03", self._predictions())

        mapping = load_entry_mapping(path, "202609")
        self.assertNotIn("axis", mapping.columns)
        self.assertNotIn("prediction_score", mapping.columns)
        row = mapping.row_for(3, 1)
        wb = load_workbook(path, data_only=True)
        ws = wb["202609"]
        self.assertEqual(ws.cell(row, mapping.columns["venue"]).value, "中山")
        self.assertEqual(ws.cell(row, mapping.columns["confidence"]).value, "A")
        wb.close()


class WriteResultsNewColumnsTest(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="io-new-columns-results-"))
        self.addCleanup(shutil.rmtree, self.tmp, True)

    def test_miss_type_written_when_column_exists(self):
        path = self.tmp / "entry_with_columns.xlsx"
        _make_entry_workbook(path, with_new_columns=True)
        results = [
            {
                "number": 1,
                "ticket_count": 6,
                "result": {
                    "trifecta": "1-2-4",
                    "status": "ハズレ",
                    "payout": 0,
                    "primary_miss_reason": "second_place_miss",
                    "secondary_miss_reasons": ["scenario_miss"],
                },
            }
        ]
        write_results(path, "2026-09-03", results)

        mapping = load_entry_mapping(path, "202609")
        row = mapping.row_for(3, 1)
        wb = load_workbook(path, data_only=True)
        ws = wb["202609"]
        self.assertEqual(ws.cell(row, mapping.columns["miss_type"]).value, "2着外れ")
        wb.close()

    def test_hit_writes_dash(self):
        path = self.tmp / "entry_with_columns_hit.xlsx"
        _make_entry_workbook(path, with_new_columns=True)
        results = [
            {
                "number": 1,
                "ticket_count": 6,
                "result": {"trifecta": "1-2-3", "status": "的中", "payout": 900},
            }
        ]
        write_results(path, "2026-09-03", results)

        mapping = load_entry_mapping(path, "202609")
        row = mapping.row_for(3, 1)
        wb = load_workbook(path, data_only=True)
        ws = wb["202609"]
        self.assertEqual(ws.cell(row, mapping.columns["miss_type"]).value, "-")
        wb.close()

    def test_no_crash_when_column_absent(self):
        path = self.tmp / "entry_without_columns.xlsx"
        _make_entry_workbook(path, with_new_columns=False)
        results = [
            {
                "number": 1,
                "ticket_count": 6,
                "result": {
                    "trifecta": "1-2-4",
                    "status": "ハズレ",
                    "payout": 0,
                    "primary_miss_reason": "axis_miss",
                },
            }
        ]
        write_results(path, "2026-09-03", results)  # 例外が出ないこと
        mapping = load_entry_mapping(path, "202609")
        self.assertNotIn("miss_type", mapping.columns)
        row = mapping.row_for(3, 1)
        wb = load_workbook(path, data_only=True)
        ws = wb["202609"]
        self.assertEqual(ws.cell(row, mapping.columns["result_trifecta"]).value, "1-2-4")
        wb.close()


class MappingAliasTest(unittest.TestCase):
    def test_new_headers_detected_without_breaking_required_columns(self):
        tmp = Path(tempfile.mkdtemp(prefix="io-mapping-"))
        try:
            path = tmp / "entry.xlsx"
            _make_entry_workbook(path, with_new_columns=True)
            wb = load_workbook(path)
            mapping = inspect_entry_sheet(wb["202609"])
            wb.close()
            self.assertEqual(mapping.columns["axis"], 15)
            self.assertEqual(mapping.columns["prediction_score"], 16)
            self.assertEqual(mapping.columns["miss_type"], 17)
            for key in (
                "pred_num",
                "target",
                "confidence",
                "venue",
                "race",
                "close_time",
                "main",
                "cover",
                "explanation",
            ):
                self.assertIn(key, mapping.columns)
            # 既存必須列の列番号は列追加前と変わらない（A〜N内に収まる）
            self.assertLessEqual(mapping.columns["explanation"], 14)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
