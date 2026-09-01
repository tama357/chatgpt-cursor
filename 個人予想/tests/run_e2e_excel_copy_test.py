#!/usr/bin/env python3
"""実Excelコピーを使った1日分通しテスト（元ファイルは変更しない）。"""

from __future__ import annotations

import hashlib
import importlib.util
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TEST_DATE = "2026-09-01"
SHEET = "202609"
TEST_EXCEL = ROOT / "excel" / "_e2e_test"
TEST_DATA = ROOT / "data" / "_e2e_test"

ORIGINALS = {
    "keiba_entry": ROOT / "excel" / "競馬_予想記入シート_2026年9月.xlsx",
    "keiba_summary": ROOT / "excel" / "競馬_予想集計シート_2026年9月.xlsx",
    "kyotei_entry": ROOT / "excel" / "競艇_予想記入シート_2026年9月.xlsx",
    "kyotei_summary": ROOT / "excel" / "競艇_予想集計シート_2026年9月.xlsx",
}


def load_workflow():
    path = ROOT / "tools" / "workflow.py"
    spec = importlib.util.spec_from_file_location("personal_workflow", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def setup_copies() -> dict[str, Path]:
    if TEST_EXCEL.exists():
        shutil.rmtree(TEST_EXCEL)
    if TEST_DATA.exists():
        shutil.rmtree(TEST_DATA)
    TEST_EXCEL.mkdir(parents=True)
    TEST_DATA.mkdir(parents=True)
    copies: dict[str, Path] = {}
    for key, src in ORIGINALS.items():
        dst = TEST_EXCEL / src.name
        shutil.copy2(src, dst)
        copies[key] = dst
    return copies


def patch_workflow(workflow, copies: dict[str, Path]) -> None:
    def test_workbooks(_base: Path) -> dict[str, Path]:
        return dict(copies)

    def test_state_path(sport: str) -> Path:
        return TEST_DATA / sport / "state.json"

    workflow.ensure_workbooks = test_workbooks  # type: ignore[method-assign]
    workflow.state_path = test_state_path  # type: ignore[method-assign]


def snapshot_formulas(path: Path, sheet: str, rows: range, cols: range) -> dict[tuple[int, int], str]:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]
    out: dict[tuple[int, int], str] = {}
    for r in rows:
        for c in cols:
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                out[(r, c)] = v
    wb.close()
    return out


def snapshot_merges_and_dv(path: Path, sheet: str) -> tuple[int, int]:
    wb = load_workbook(path)
    ws = wb[sheet]
    merges = len(list(ws.merged_cells.ranges))
    dvs = len(ws.data_validations.dataValidation)
    wb.close()
    return merges, dvs


def read_entry_rows(path: Path, start: int = 3, end: int = 7) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in range(start, end + 1):
        rows.append({chr(64 + c): ws.cell(r, c).value for c in range(1, 15)})
    wb.close()
    return rows


def read_summary_pt(path: Path) -> dict[str, list]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]
    data = {
        "row12_P-T": [ws.cell(12, c).value for c in range(16, 21)],
        "row13_P-T": [ws.cell(13, c).value for c in range(16, 21)],
        "row14_P-T": [ws.cell(14, c).value for c in range(16, 21)],
        "当日的中率": ws.cell(12, 9).value,
        "当日回収率": ws.cell(12, 10).value,
        "累計的中率": ws.cell(12, 3).value,
        "累計回収率": ws.cell(12, 4).value,
    }
    wb.close()
    return data


def cols_c_to_k_only_updated(before: list[dict], after: list[dict]) -> list[str]:
    errors = []
    for i, (b, a) in enumerate(zip(before, after), start=3):
        for col in "AB":
            if b.get(col) != a.get(col):
                errors.append(f"行{i} {col}列が変化: {b.get(col)!r} -> {a.get(col)!r}")
        for col in "LMN":
            if b.get(col) != a.get(col) and a.get(col) not in (None, "未実施"):
                pass  # 予想段階ではL-Nは未更新が正常
        for col in "CDEFGHIJK":
            if b.get(col) != a.get(col):
                if a.get(col) in (None, "") and b.get(col) in (None, ""):
                    continue
                # updated expected
                continue
        if all(a.get(c) in (None, "") for c in "CDEFGHIJK"):
            errors.append(f"行{i} C-K列が空のまま")
    return errors


def main() -> int:
    workflow = load_workflow()
    original_hashes = {k: file_hash(p) for k, p in ORIGINALS.items()}
    copies = setup_copies()
    patch_workflow(workflow, copies)
    for sport in ("keiba", "kyotei"):
        dest = ROOT / "data" / "races" / sport / f"{TEST_DATE}.json"
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(ROOT / "examples" / f"{sport}_races.sample.json", dest)

    before_keiba_entry = read_entry_rows(copies["keiba_entry"])
    before_kyotei_entry = read_entry_rows(copies["kyotei_entry"])
    formula_before_kb = snapshot_formulas(
        copies["keiba_summary"], SHEET, range(12, 15), range(2, 16)
    )
    formula_before_kr = snapshot_formulas(
        copies["kyotei_summary"], SHEET, range(12, 15), range(2, 16)
    )
    merge_before_kb = snapshot_merges_and_dv(copies["keiba_entry"], SHEET)
    merge_before_kr = snapshot_merges_and_dv(copies["kyotei_entry"], SHEET)

    report: dict[str, object] = {"date": TEST_DATE, "checks": []}

    # 1. predict-all
    kb_pred = workflow.run_predict("keiba", TEST_DATE, force=True, sync_drive=False)
    kr_pred = workflow.run_predict("kyotei", TEST_DATE, force=True, sync_drive=False)
    report["predict_keiba_head"] = kb_pred.splitlines()[:5]
    report["predict_kyotei_head"] = kr_pred.splitlines()[:5]

    after_keiba_entry = read_entry_rows(copies["keiba_entry"])
    after_kyotei_entry = read_entry_rows(copies["kyotei_entry"])

    keiba_ab_ok = all(
        after_keiba_entry[i]["A"] == before_keiba_entry[i]["A"]
        and after_keiba_entry[i]["B"] == before_keiba_entry[i]["B"]
        for i in range(5)
    )
    kyotei_ab_ok = all(
        after_kyotei_entry[i]["A"] == before_kyotei_entry[i]["A"]
        and after_kyotei_entry[i]["B"] == before_kyotei_entry[i]["B"]
        for i in range(5)
    )
    keiba_ck_updated = all(
        any(after_keiba_entry[i][c] not in (None, "") for c in "CDEFGHIJK")
        for i in range(3)
    )
    kyotei_ck_updated = all(
        any(after_kyotei_entry[i][c] not in (None, "") for c in "CDEFGHIJK")
        for i in range(3)
    )
    merge_after_kb = snapshot_merges_and_dv(copies["keiba_entry"], SHEET)
    merge_after_kr = snapshot_merges_and_dv(copies["kyotei_entry"], SHEET)

    report["checks"].append(
        {
            "step": "predict-all",
            "keiba_AB_preserved": keiba_ab_ok,
            "kyotei_AB_preserved": kyotei_ab_ok,
            "keiba_C-K_updated_rows3-5": keiba_ck_updated,
            "kyotei_C-K_updated_rows3-5": kyotei_ck_updated,
            "keiba_merges_unchanged": merge_before_kb[0] == merge_after_kb[0],
            "kyotei_merges_unchanged": merge_before_kr[0] == merge_after_kr[0],
            "keiba_dv_unchanged": merge_before_kb[1] == merge_after_kb[1],
            "kyotei_dv_unchanged": merge_before_kr[1] == merge_after_kr[1],
            "keiba_rows_3_7": after_keiba_entry,
            "kyotei_rows_3_7": after_kyotei_entry,
        }
    )

    # 2. apply-results + results
    workflow.apply_results_from_file(
        "keiba", TEST_DATE, ROOT / "examples" / "keiba_results.sample.json", sync_drive=False
    )
    workflow.apply_results_from_file(
        "kyotei", TEST_DATE, ROOT / "examples" / "kyotei_results.sample.json", sync_drive=False
    )

    entry_after_res_kb = read_entry_rows(copies["keiba_entry"])
    entry_after_res_kr = read_entry_rows(copies["kyotei_entry"])
    summary_kb = read_summary_pt(copies["keiba_summary"])
    summary_kr = read_summary_pt(copies["kyotei_summary"])
    formula_after_kb = snapshot_formulas(
        copies["keiba_summary"], SHEET, range(12, 15), range(2, 16)
    )
    formula_after_kr = snapshot_formulas(
        copies["kyotei_summary"], SHEET, range(12, 15), range(2, 16)
    )

    # formula errors check
    wb = load_workbook(copies["keiba_summary"])
    formula_errors_kb = wb[SHEET].calculate_dimension()  # placeholder
    wb.close()

    def has_formula_errors(path: Path) -> bool:
        wb = load_workbook(path)
        ws = wb[SHEET]
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=20):
            for cell in row:
                if cell.data_type == "e":
                    wb.close()
                    return True
        wb.close()
        return False

    ln_kb = [entry_after_res_kb[i]["L"] for i in range(3)]
    ln_kr = [entry_after_res_kr[i]["L"] for i in range(3)]

    report["checks"].append(
        {
            "step": "apply-results",
            "keiba_L-N_sample": [{k: entry_after_res_kb[i][k] for k in "LMN"} for i in range(3)],
            "kyotei_L-N_sample": [{k: entry_after_res_kr[i][k] for k in "LMN"} for i in range(3)],
            "keiba_summary_PT": summary_kb,
            "kyotei_summary_PT": summary_kr,
            "keiba_B-O_formulas_unchanged": formula_before_kb == formula_after_kb,
            "kyotei_B-O_formulas_unchanged": formula_before_kr == formula_after_kr,
            "keiba_formula_errors": has_formula_errors(copies["keiba_summary"]),
            "kyotei_formula_errors": has_formula_errors(copies["kyotei_summary"]),
        }
    )

    # 3. idempotency
    dup_kb = workflow.run_predict("keiba", TEST_DATE, force=False, sync_drive=False)
    dup_kr = workflow.run_predict("kyotei", TEST_DATE, force=False, sync_drive=False)
    forced_kb = workflow.run_predict("keiba", TEST_DATE, force=True, sync_drive=False)

    report["checks"].append(
        {
            "step": "idempotency",
            "dup_keiba_blocked": "二重登録防止" in dup_kb,
            "dup_kyotei_blocked": "二重登録防止" in dup_kr,
            "force_keiba_ran": "選定レース数" in forced_kb or "予想報告" in forced_kb,
        }
    )

    # 4. originals unchanged
    unchanged = all(file_hash(ORIGINALS[k]) == original_hashes[k] for k in ORIGINALS)
    report["checks"].append({"step": "isolation", "originals_unchanged": unchanged})

    # Chatwork: grep workflow source
    src = (ROOT / "tools" / "workflow.py").read_text()
    report["checks"].append(
        {
            "step": "chatwork",
            "workflow_calls_chatwork": "send_chatwork" in src or "chatwork_request" in src,
        }
    )

    out = ROOT / "tests" / "e2e_excel_test_report.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    failed = []
    for chk in report["checks"]:
        if chk["step"] == "predict-all":
            for k, v in chk.items():
                if k.endswith("_unchanged") or k.endswith("_preserved") or k.endswith("updated_rows3-5"):
                    if v is False:
                        failed.append(f"predict-all:{k}")
        if chk["step"] == "apply-results":
            if not chk["keiba_B-O_formulas_unchanged"]:
                failed.append("summary formulas keiba changed")
            if not chk["kyotei_B-O_formulas_unchanged"]:
                failed.append("summary formulas kyotei changed")
            if chk["keiba_formula_errors"] or chk["kyotei_formula_errors"]:
                failed.append("formula errors detected")
            keiba_ln = chk.get("keiba_L-N_sample") or []
            kyotei_ln = chk.get("kyotei_L-N_sample") or []
            if not keiba_ln or any(not row.get("L") for row in keiba_ln):
                failed.append("keiba results L column empty")
            if not kyotei_ln or any(not row.get("L") for row in kyotei_ln):
                failed.append("kyotei results L column empty")
            kyotei_pt = chk.get("kyotei_summary_PT") or {}
            kyotei_status = kyotei_pt.get("row12_P-T") or []
            if not any(s in {"的中", "ハズレ"} for s in kyotei_status):
                failed.append("kyotei summary P-T not updated")
        if chk["step"] == "idempotency":
            if not chk["dup_keiba_blocked"] or not chk["dup_kyotei_blocked"]:
                failed.append("idempotency failed")
        if chk["step"] == "isolation" and not chk["originals_unchanged"]:
            failed.append("originals modified")
    if failed:
        print("FAILED:", failed, file=sys.stderr)
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
