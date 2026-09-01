from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from common.constants import EXCEL_FILENAMES, MONTH_SHEETS

ENTRY_HEADER_ALIASES = {
    "date": ("日付",),
    "pred_num": ("予想番号",),
    "target": ("狙い",),
    "confidence": ("自信度",),
    "venue": ("競馬場", "競輪場", "競艇場"),
    "race": ("R", "r"),
    "close_time": ("締切時刻",),
    "main": ("本線",),
    "cover": ("抑え",),
    "total_points": ("合計点数",),
    "explanation": ("解説",),
    "result_trifecta": ("結果 3連単", "結果3連単"),
    "payout": ("払戻金",),
    "status": ("結果",),
}

SUMMARY_RACE_HEADERS = ("1本目", "2本目", "3本目", "4本目", "5本目")


@dataclass
class EntrySheetMapping:
    header_row: int
    data_start_row: int
    rows_per_day: int = 5
    columns: dict[str, int] = field(default_factory=dict)

    def row_for(self, day: int, pred_num: int) -> int:
        return self.data_start_row + (day - 1) * self.rows_per_day + (pred_num - 1)


@dataclass
class SummarySheetMapping:
    header_row: int
    data_start_row: int
    rows_per_day: int = 3
    label_col: int = 15
    race_cols: list[int] = field(default_factory=list)

    def day_start_row(self, day: int) -> int:
        return self.data_start_row + (day - 1) * self.rows_per_day


def month_sheet_name(date_str: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return f"{dt.year}{dt.month:02d}"


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).replace("\n", ""))


def _find_col(headers: dict[int, str], aliases: tuple[str, ...]) -> int | None:
    normalized_aliases = {_norm_header(a) for a in aliases}
    for col, title in headers.items():
        if _norm_header(title) in normalized_aliases:
            return col
    for col, title in headers.items():
        norm = _norm_header(title)
        if any(a in norm for a in normalized_aliases):
            return col
    return None


def inspect_entry_sheet(ws: Worksheet) -> EntrySheetMapping:
    header_row = None
    for row in range(1, 20):
        headers = {col: ws.cell(row, col).value for col in range(1, 20)}
        if _find_col(headers, ENTRY_HEADER_ALIASES["date"]) and _find_col(
            headers, ENTRY_HEADER_ALIASES["pred_num"]
        ):
            header_row = row
            break
    if header_row is None:
        raise ValueError(f"予想記入シート '{ws.title}' のヘッダ行が見つかりません")

    headers = {col: ws.cell(header_row, col).value for col in range(1, 20)}
    columns: dict[str, int] = {}
    for key, aliases in ENTRY_HEADER_ALIASES.items():
        col = _find_col(headers, aliases)
        if col is not None:
            columns[key] = col

    required = ["pred_num", "target", "confidence", "venue", "race", "close_time", "main", "cover", "explanation"]
    missing = [k for k in required if k not in columns]
    if missing:
        raise ValueError(f"予想記入シート '{ws.title}' に必要列がありません: {missing}")

    data_start = header_row + 1
    rows_per_day = 5
    for probe in range(5):
        if ws.cell(data_start + probe, columns["pred_num"]).value == 1:
            first_one = data_start + probe
            break
    else:
        first_one = data_start

    for diff in (5, 4, 6, 3):
        if ws.cell(first_one + diff, columns["pred_num"]).value == 1:
            rows_per_day = diff
            break

    return EntrySheetMapping(
        header_row=header_row,
        data_start_row=first_one,
        rows_per_day=rows_per_day,
        columns=columns,
    )


def inspect_summary_sheet(ws: Worksheet) -> SummarySheetMapping:
    header_row = None
    race_cols: list[int] = []
    for row in range(1, 30):
        headers = {col: ws.cell(row, col).value for col in range(1, 25)}
        texts = [_norm_header(v) for v in headers.values()]
        if "日付" in texts and any(h in texts for h in SUMMARY_RACE_HEADERS):
            header_row = row
            for col, val in headers.items():
                if _norm_header(val) in SUMMARY_RACE_HEADERS:
                    race_cols.append(col)
            break
    if header_row is None or len(race_cols) < 5:
        raise ValueError(f"集計シート '{ws.title}' のヘッダ行が見つかりません")

    race_cols = sorted(race_cols)[:5]
    label_col = race_cols[0] - 1
    data_start = header_row + 1

    rows_per_day = 3
    for diff in (3, 4, 2):
        val = ws.cell(data_start + diff, 2).value
        if val is not None and hasattr(val, "day"):
            rows_per_day = diff
            break

    return SummarySheetMapping(
        header_row=header_row,
        data_start_row=data_start,
        rows_per_day=rows_per_day,
        label_col=label_col,
        race_cols=race_cols,
    )


def inspect_workbook(path: Path, *, kind: str) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, Any] = {"path": str(path), "kind": kind, "sheets": {}}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if kind == "entry":
            mapping = inspect_entry_sheet(ws)
            result["sheets"][sheet_name] = {
                "header_row": mapping.header_row,
                "data_start_row": mapping.data_start_row,
                "rows_per_day": mapping.rows_per_day,
                "columns": mapping.columns,
            }
        else:
            mapping = inspect_summary_sheet(ws)
            result["sheets"][sheet_name] = {
                "header_row": mapping.header_row,
                "data_start_row": mapping.data_start_row,
                "rows_per_day": mapping.rows_per_day,
                "label_col": mapping.label_col,
                "race_cols": mapping.race_cols,
            }
    wb.close()
    return result


def load_entry_mapping(path: Path, sheet_name: str) -> EntrySheetMapping:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"シート {sheet_name} がありません: {path.name}")
    mapping = inspect_entry_sheet(wb[sheet_name])
    wb.close()
    return mapping


def load_summary_mapping(path: Path, sheet_name: str) -> SummarySheetMapping:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError(f"シート {sheet_name} がありません: {path.name}")
    mapping = inspect_summary_sheet(wb[sheet_name])
    wb.close()
    return mapping


def write_mapping_cache(base_dir: Path) -> Path:
    excel_dir = base_dir / "excel"
    files = {key: excel_dir / name for key, name in EXCEL_FILENAMES.items()}
    cache: dict[str, Any] = {}
    for key, path in files.items():
        kind = "entry" if "entry" in key else "summary"
        cache[key] = inspect_workbook(path, kind=kind)
    out = excel_dir / "sheet_mapping.json"
    out.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return out


def safe_write(ws: Worksheet, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return
    cell.value = value
