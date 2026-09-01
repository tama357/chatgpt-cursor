from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from excel.layout import (
    BLOCK_HEIGHT,
    COL_AXIS,
    COL_CLOSE,
    COL_CONFIDENCE,
    COL_COVER,
    COL_MAIN,
    COL_PAYOUT,
    COL_POINTS,
    COL_PRED_NUM,
    COL_PRED_SCORE,
    COL_RACE,
    COL_RATIONALE,
    COL_RESULT,
    COL_RISKS,
    COL_SCENARIO,
    COL_SKIP,
    COL_SPORT,
    COL_STATUS,
    COL_TARGET,
    COL_TOTAL,
    COL_VENUE,
    MAX_RACES,
    SUMMARY_RACE_COLS,
    SUMMARY_ROWS_PER_DAY,
    blocks,
    main_rows,
    summary_day_start_row,
    summary_tab_name,
)


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2EFDA")


def _set_header(ws, row: int, headers: list[tuple[int, str]]) -> None:
    for col, title in headers:
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_entry_template_ws(ws, sport_label: str) -> None:
    ws.title = "テンプレ"
    headers = [
        (COL_PRED_NUM, "予想#"),
        (COL_SPORT, "競技"),
        (COL_VENUE, "開催場"),
        (COL_RACE, "R"),
        (COL_CLOSE, "締切"),
        (COL_TARGET, "狙い"),
        (COL_PRED_SCORE, "予想しやすさ"),
        (COL_CONFIDENCE, "自信度"),
        (COL_AXIS, "軸"),
        (COL_MAIN, "本線"),
        (COL_COVER, "抑え"),
        (COL_TOTAL, "合計点数"),
        (COL_RATIONALE, "予想根拠"),
        (COL_SCENARIO, "想定展開"),
        (COL_RISKS, "主なリスク"),
        (COL_SKIP, "見送り"),
        (COL_RESULT, "三連単結果"),
        (COL_PAYOUT, "払戻金"),
        (COL_STATUS, "的中/ハズレ"),
        (COL_POINTS, "購入点数"),
    ]
    _set_header(ws, 1, headers)
    target_dv = DataValidation(type="list", formula1='"鉄板,中穴,大穴"', allow_blank=True)
    conf_dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
    status_dv = DataValidation(type="list", formula1='"的中,ハズレ,未実施"', allow_blank=True)
    ws.add_data_validation(target_dv)
    ws.add_data_validation(conf_dv)
    ws.add_data_validation(status_dv)

    for block in blocks():
        row = block.main_row
        ws.cell(row=row, column=COL_PRED_NUM, value=block.index)
        ws.cell(row=row, column=COL_SPORT, value=sport_label)
        for col in range(1, 21):
            cell = ws.cell(row=row, column=col)
            if col <= COL_SKIP:
                cell.fill = INPUT_FILL
        target_dv.add(ws.cell(row=row, column=COL_TARGET))
        conf_dv.add(ws.cell(row=row, column=COL_CONFIDENCE))
        status_dv.add(ws.cell(row=row, column=COL_STATUS))
        note_row = row + 1
        ws.cell(row=note_row, column=COL_RATIONALE, value="※このブロックは予想1件分。数式列は触らない。")
        for r in range(row + 1, row + BLOCK_HEIGHT):
            ws.row_dimensions[r].hidden = False

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["M"].width = 40
    ws.column_dimensions["N"].width = 30
    ws.column_dimensions["O"].width = 30


def _build_summary_template_ws(ws, year: int, month: int) -> None:
    tab = summary_tab_name(year, month)
    ws.title = tab
    ws.cell(row=1, column=1, value="日付")
    for idx, col in enumerate(SUMMARY_RACE_COLS, start=1):
        ws.cell(row=1, column=col, value=f"R{idx}")
    ws.cell(row=1, column=21, value="備考")

    for day in range(1, 32):
        start = summary_day_start_row(day)
        ws.cell(row=start, column=1, value=f"{month}/{day}")
        ws.cell(row=start + 1, column=1, value="的中額")
        ws.cell(row=start + 2, column=1, value="購入点数")
        for offset in range(SUMMARY_ROWS_PER_DAY):
            for col in SUMMARY_RACE_COLS:
                ws.cell(row=start + offset, column=col).fill = INPUT_FILL
        for col in range(2, 16):
            for offset in range(SUMMARY_ROWS_PER_DAY):
                cell = ws.cell(row=start + offset, column=col)
                cell.fill = FORMULA_FILL
                cell.value = f"=0"


def create_entry_workbook(path: Path, sport_label: str) -> None:
    wb = Workbook()
    _build_entry_template_ws(wb.active, sport_label)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def create_summary_workbook(path: Path, year: int, month: int) -> None:
    wb = Workbook()
    _build_summary_template_ws(wb.active, year, month)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def ensure_workbooks(base_dir: Path) -> dict[str, Path]:
    excel_dir = base_dir / "excel"
    files = {
        "keiba_entry": excel_dir / "競馬_予想記入シート_2026年9月.xlsx",
        "keiba_summary": excel_dir / "競馬_予想集計シート_2026年9月.xlsx",
        "keirin_entry": excel_dir / "競輪_個人_予想記入シート_2026年9月.xlsx",
        "keirin_summary": excel_dir / "競輪_個人_予想集計シート_2026年9月.xlsx",
    }
    if not files["keiba_entry"].exists():
        create_entry_workbook(files["keiba_entry"], "競馬")
    if not files["keiba_summary"].exists():
        create_summary_workbook(files["keiba_summary"], 2026, 9)
    if not files["keirin_entry"].exists():
        create_entry_workbook(files["keirin_entry"], "競輪")
    if not files["keirin_summary"].exists():
        create_summary_workbook(files["keirin_summary"], 2026, 9)
    return files


def create_all_templates(base_dir: Path) -> None:
    ensure_workbooks(base_dir)
