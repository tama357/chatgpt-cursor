from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common.constants import MISS_TYPE_MAP, MISS_TYPE_NONE
from common.tickets import expand_tickets
from excel.mapping import (
    load_entry_mapping,
    load_summary_mapping,
    month_sheet_name,
    safe_write,
)


def _format_explanation(pred: dict[str, Any]) -> str:
    base = pred.get("rationale") or pred.get("explanation") or ""
    score = pred.get("prediction_score")
    if score is not None:
        prefix = f"[予想しやすさ{score}]"
        if prefix not in base:
            base = f"{prefix} {base}".strip()
    scenario = pred.get("scenario")
    risks = pred.get("risks")
    parts = [base]
    if scenario:
        parts.append(f"想定:{scenario}")
    if risks:
        parts.append(f"リスク:{risks}")
    return " / ".join(p for p in parts if p)


def _miss_type_label(result: dict[str, Any]) -> str:
    """外れ型（表示用）。state.jsonのprimary_miss_reasonを1対1で日本語化するだけで、
    secondary_miss_reasonsを含む詳細情報はstate.json側にそのまま残す。"""
    if result.get("status") != "ハズレ":
        return MISS_TYPE_NONE
    reason = result.get("primary_miss_reason")
    return MISS_TYPE_MAP.get(reason, MISS_TYPE_NONE)


def write_predictions(
    entry_path: Path,
    date_str: str,
    predictions: list[dict[str, Any]],
) -> str:
    sheet = month_sheet_name(date_str)
    mapping = load_entry_mapping(entry_path, sheet)
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    day = dt.day
    cols = mapping.columns

    wb = load_workbook(entry_path)
    ws = wb[sheet]

    for pred in predictions:
        number = pred.get("number")
        if not isinstance(number, int) or not 1 <= number <= 5:
            continue
        row = mapping.row_for(day, number)
        tickets = pred.get("tickets", [])
        expanded = expand_tickets(tickets) if tickets else []
        main_picks = [t.compact for t in expanded if t.kind == "本線"]
        cover_picks = [t.compact for t in expanded if t.kind == "抑え"]
        total = sum(len(t.combinations) for t in expanded)

        writes = {
            "target": pred.get("target"),
            "confidence": pred.get("confidence"),
            "venue": pred.get("venue"),
            "race": pred.get("race"),
            "close_time": pred.get("close_time"),
            "main": " ".join(main_picks),
            "cover": " ".join(cover_picks),
            "explanation": _format_explanation(pred),
            "axis": pred.get("axis"),
            "prediction_score": pred.get("prediction_score"),
        }
        if "total_points" in cols:
            writes["total_points"] = total

        for key, val in writes.items():
            col = cols.get(key)
            if col:
                safe_write(ws, row, col, val)

    wb.save(entry_path)
    wb.close()
    return f"記入完了: {entry_path.name} / シート {sheet} / {len(predictions)}レース（{day}日・予想行のみ）"


def write_results(
    entry_path: Path,
    date_str: str,
    results: list[dict[str, Any]],
) -> str:
    sheet = month_sheet_name(date_str)
    mapping = load_entry_mapping(entry_path, sheet)
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    day = dt.day
    cols = mapping.columns

    wb = load_workbook(entry_path)
    ws = wb[sheet]

    for item in results:
        number = item.get("number")
        if not isinstance(number, int) or not 1 <= number <= 5:
            continue
        row = mapping.row_for(day, number)
        result = item.get("result", item)
        if "result_trifecta" in cols:
            safe_write(ws, row, cols["result_trifecta"], result.get("trifecta"))
        if "payout" in cols:
            safe_write(ws, row, cols["payout"], result.get("payout"))
        if "status" in cols:
            safe_write(ws, row, cols["status"], result.get("status"))
        if "total_points" in cols:
            safe_write(ws, row, cols["total_points"], item.get("ticket_count"))
        if "miss_type" in cols:
            safe_write(ws, row, cols["miss_type"], _miss_type_label(result))

    wb.save(entry_path)
    wb.close()
    return f"結果記入完了: {entry_path.name} / シート {sheet}"


def write_summary(
    summary_path: Path,
    date_str: str,
    results: list[dict[str, Any]],
) -> str:
    sheet = month_sheet_name(date_str)
    mapping = load_summary_mapping(summary_path, sheet)
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    day = dt.day
    start = mapping.day_start_row(day)

    wb = load_workbook(summary_path)
    ws = wb[sheet]

    for idx, item in enumerate(results[:5]):
        if idx >= len(mapping.race_cols):
            break
        col = mapping.race_cols[idx]
        result = item.get("result", item)
        status = result.get("status", "")
        payout = result.get("payout", 0) if status == "的中" else 0
        points = item.get("ticket_count") or result.get("points", 0)
        safe_write(ws, start, col, status)
        safe_write(ws, start + 1, col, payout)
        safe_write(ws, start + 2, col, points)

    wb.save(summary_path)
    wb.close()
    return f"集計シート更新: {summary_path.name} / シート {sheet} / {day}日 / P-T列相当"
